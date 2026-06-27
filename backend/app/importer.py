from __future__ import annotations

import hashlib
import json
from datetime import date, datetime
from decimal import Decimal
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from sqlalchemy import text
from sqlalchemy.engine import Connection

from .config import DOCS_DIR


BUSINESS_SHEET_INDEX = 1
HEADER_ROW = 3
DATA_START_ROW = 5


def _json_default(value: Any) -> str | float | int | None:
    if isinstance(value, (datetime, date)):
        return value.isoformat()
    if isinstance(value, Decimal):
        return float(value)
    return str(value) if value is not None else None


def _as_date(value: Any) -> date | None:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    return None


def _as_text(value: Any) -> str | None:
    if value is None:
        return None
    text_value = str(value).strip()
    return text_value or None


def _as_decimal(value: Any) -> float | None:
    if value is None or value == "":
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def _row_value(row: tuple[Any, ...], position: int) -> Any:
    index = position - 1
    return row[index] if 0 <= index < len(row) else None


def _execute_scalar(conn: Connection, sql: str, params: dict[str, Any]) -> int:
    result = conn.execute(text(sql), params)
    if result.lastrowid:
        return int(result.lastrowid)
    return int(conn.execute(text("SELECT LAST_INSERT_ID()")).scalar() or 0)


def _reset_business_data(conn: Connection) -> None:
    tables = [
        "sales_receipt",
        "sales_invoice",
        "sales_contract",
        "purchase_payment",
        "finance_invoice_check",
        "warehouse_entry",
        "purchase_invoice",
        "purchase_contract",
        "delivery_record",
        "purchase_info",
        "order_line",
        "sales_order",
        "project",
        "ledger_raw_row",
        "backup_record",
        "operation_log",
        "erp_user",
        "import_batch",
    ]
    conn.execute(text("SET FOREIGN_KEY_CHECKS=0"))
    for table in tables:
        conn.execute(text(f"TRUNCATE TABLE {table}"))
    conn.execute(text("SET FOREIGN_KEY_CHECKS=1"))


def import_excel(conn: Connection, reset: bool = True) -> dict[str, Any]:
    if reset:
        _reset_business_data(conn)

    workbook_path = next(DOCS_DIR.glob("2026*.xlsx"))
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    worksheet = workbook.worksheets[BUSINESS_SHEET_INDEX]
    headers = list(next(worksheet.iter_rows(min_row=HEADER_ROW, max_row=HEADER_ROW, values_only=True)))

    batch_id = _execute_scalar(
        conn,
        """
        INSERT INTO import_batch
          (source_file_name, source_sheet_name, header_row_no, data_start_row_no, status)
        VALUES
          (:file_name, :sheet_name, :header_row_no, :data_start_row_no, 'parsing')
        """,
        {
            "file_name": workbook_path.name,
            "sheet_name": worksheet.title,
            "header_row_no": HEADER_ROW,
            "data_start_row_no": DATA_START_ROW,
        },
    )

    success_rows = 0
    failed_rows = 0

    for excel_row_no, row in enumerate(
        worksheet.iter_rows(min_row=DATA_START_ROW, values_only=True),
        start=DATA_START_ROW,
    ):
        project_code = _as_text(_row_value(row, 2))
        order_no = _as_text(_row_value(row, 13))
        if not project_code or not order_no:
            continue

        row_dict = {
            str(headers[i] or f"column_{i + 1}"): _json_default(value)
            for i, value in enumerate(row)
        }
        raw_json = json.dumps(row_dict, ensure_ascii=False, default=_json_default)
        row_hash = hashlib.sha256(raw_json.encode("utf-8")).hexdigest()

        try:
            raw_row_id = _execute_scalar(
                conn,
                """
                INSERT INTO ledger_raw_row
                  (import_batch_id, excel_row_no, row_hash, raw_json, parse_status, project_code, order_no)
                VALUES
                  (:batch_id, :excel_row_no, :row_hash, CAST(:raw_json AS JSON), 'success', :project_code, :order_no)
                ON DUPLICATE KEY UPDATE id = LAST_INSERT_ID(id), parse_status = 'success'
                """,
                {
                    "batch_id": batch_id,
                    "excel_row_no": excel_row_no,
                    "row_hash": row_hash,
                    "raw_json": raw_json,
                    "project_code": project_code,
                    "order_no": order_no,
                },
            )

            project_id = _execute_scalar(
                conn,
                """
                INSERT INTO project
                  (project_code, project_name, department, branch_company, account_manager,
                   team_level3_name, customer_unit_name, end_user_name, regional_platform)
                VALUES
                  (:project_code, :project_name, :department, :branch_company, :account_manager,
                   :team_level3_name, :customer_unit_name, :end_user_name, :regional_platform)
                ON DUPLICATE KEY UPDATE
                  id = LAST_INSERT_ID(id),
                  project_name = COALESCE(VALUES(project_name), project_name),
                  department = COALESCE(VALUES(department), department),
                  branch_company = COALESCE(VALUES(branch_company), branch_company),
                  account_manager = COALESCE(VALUES(account_manager), account_manager),
                  team_level3_name = COALESCE(VALUES(team_level3_name), team_level3_name),
                  customer_unit_name = COALESCE(VALUES(customer_unit_name), customer_unit_name),
                  end_user_name = COALESCE(VALUES(end_user_name), end_user_name),
                  regional_platform = COALESCE(VALUES(regional_platform), regional_platform)
                """,
                {
                    "project_code": project_code,
                    "project_name": _as_text(_row_value(row, 14)),
                    "department": _as_text(_row_value(row, 3)),
                    "branch_company": _as_text(_row_value(row, 4)),
                    "account_manager": _as_text(_row_value(row, 5)),
                    "team_level3_name": _as_text(_row_value(row, 9)),
                    "customer_unit_name": _as_text(_row_value(row, 10)),
                    "end_user_name": _as_text(_row_value(row, 11)),
                    "regional_platform": _as_text(_row_value(row, 12)),
                },
            )

            sales_order_id = _execute_scalar(
                conn,
                """
                INSERT INTO sales_order
                  (project_id, import_batch_id, source_excel_row_no, gross_net_type, order_no,
                   order_date, business_type, statistic_category, close_status)
                VALUES
                  (:project_id, :batch_id, :excel_row_no, :gross_net_type, :order_no,
                   :order_date, :business_type, :statistic_category, :close_status)
                ON DUPLICATE KEY UPDATE
                  id = LAST_INSERT_ID(id),
                  order_date = COALESCE(VALUES(order_date), order_date),
                  business_type = COALESCE(VALUES(business_type), business_type),
                  statistic_category = COALESCE(VALUES(statistic_category), statistic_category),
                  close_status = COALESCE(VALUES(close_status), close_status)
                """,
                {
                    "project_id": project_id,
                    "batch_id": batch_id,
                    "excel_row_no": excel_row_no,
                    "gross_net_type": _as_text(_row_value(row, 1)),
                    "order_no": order_no,
                    "order_date": _as_date(_row_value(row, 6)),
                    "business_type": _as_text(_row_value(row, 7)),
                    "statistic_category": _as_text(_row_value(row, 8)),
                    "close_status": _as_text(_row_value(row, 87)),
                },
            )

            order_line_id = _execute_scalar(
                conn,
                """
                INSERT INTO order_line
                  (sales_order_id, raw_row_id, source_excel_row_no, goods_name, specification_model,
                   unit_name, quantity, sales_unit_price_no_tax, sales_unit_price, revenue_no_tax, order_value)
                VALUES
                  (:sales_order_id, :raw_row_id, :excel_row_no, :goods_name, :specification_model,
                   :unit_name, :quantity, :sales_unit_price_no_tax, :sales_unit_price, :revenue_no_tax, :order_value)
                """,
                {
                    "sales_order_id": sales_order_id,
                    "raw_row_id": raw_row_id,
                    "excel_row_no": excel_row_no,
                    "goods_name": _as_text(_row_value(row, 15)),
                    "specification_model": _as_text(_row_value(row, 16)),
                    "unit_name": _as_text(_row_value(row, 17)),
                    "quantity": _as_decimal(_row_value(row, 18)),
                    "sales_unit_price_no_tax": _as_decimal(_row_value(row, 19)),
                    "sales_unit_price": _as_decimal(_row_value(row, 20)),
                    "revenue_no_tax": _as_decimal(_row_value(row, 21)),
                    "order_value": _as_decimal(_row_value(row, 22)),
                },
            )

            conn.execute(
                text(
                    """
                    INSERT INTO purchase_info
                      (order_line_id, supplier_name, purchase_unit_price_no_tax, purchase_unit_price,
                       cost_no_tax, purchase_amount)
                    VALUES
                      (:order_line_id, :supplier_name, :purchase_unit_price_no_tax, :purchase_unit_price,
                       :cost_no_tax, :purchase_amount)
                    """
                ),
                {
                    "order_line_id": order_line_id,
                    "supplier_name": _as_text(_row_value(row, 23)),
                    "purchase_unit_price_no_tax": _as_decimal(_row_value(row, 24)),
                    "purchase_unit_price": _as_decimal(_row_value(row, 25)),
                    "cost_no_tax": _as_decimal(_row_value(row, 26)),
                    "purchase_amount": _as_decimal(_row_value(row, 27)),
                },
            )

            conn.execute(
                text(
                    """
                    INSERT INTO delivery_record
                      (order_line_id, delivery_date, delivery_quantity, delivery_revenue_no_tax,
                       delivery_value, delivery_cost_no_tax, delivery_cost, pending_delivery_quantity,
                       pending_delivery_amount_no_tax, pending_delivery_amount)
                    VALUES
                      (:order_line_id, :delivery_date, :delivery_quantity, :delivery_revenue_no_tax,
                       :delivery_value, :delivery_cost_no_tax, :delivery_cost, :pending_delivery_quantity,
                       :pending_delivery_amount_no_tax, :pending_delivery_amount)
                    """
                ),
                {
                    "order_line_id": order_line_id,
                    "delivery_date": _as_date(_row_value(row, 28)),
                    "delivery_quantity": _as_decimal(_row_value(row, 29)),
                    "delivery_revenue_no_tax": _as_decimal(_row_value(row, 30)),
                    "delivery_value": _as_decimal(_row_value(row, 31)),
                    "delivery_cost_no_tax": _as_decimal(_row_value(row, 32)),
                    "delivery_cost": _as_decimal(_row_value(row, 33)),
                    "pending_delivery_quantity": _as_decimal(_row_value(row, 34)),
                    "pending_delivery_amount_no_tax": _as_decimal(_row_value(row, 35)),
                    "pending_delivery_amount": _as_decimal(_row_value(row, 36)),
                },
            )

            conn.execute(
                text(
                    """
                    INSERT INTO purchase_contract
                      (order_line_id, purchase_contract_no, payment_terms, performance_period, signed_amount, unsigned_amount)
                    VALUES
                      (:order_line_id, :purchase_contract_no, :payment_terms, :performance_period, :signed_amount, :unsigned_amount)
                    """
                ),
                {
                    "order_line_id": order_line_id,
                    "purchase_contract_no": _as_text(_row_value(row, 37)),
                    "payment_terms": _as_text(_row_value(row, 38)),
                    "performance_period": _as_text(_row_value(row, 39)),
                    "signed_amount": _as_decimal(_row_value(row, 40)),
                    "unsigned_amount": _as_decimal(_row_value(row, 41)),
                },
            )

            _insert_phase(conn, "purchase_invoice", order_line_id, 1, row, 42, 43, 44)
            _insert_phase(conn, "warehouse_entry", order_line_id, 1, row, 45, 46, 47)
            _insert_payment(conn, order_line_id, 1, _as_date(_row_value(row, 52)), _as_date(_row_value(row, 53)), _as_text(_row_value(row, 54)), _as_decimal(_row_value(row, 55)))
            _insert_payment(conn, order_line_id, 2, None, _as_date(_row_value(row, 56)), _as_text(_row_value(row, 57)), _as_decimal(_row_value(row, 58)))

            conn.execute(
                text(
                    """
                    INSERT INTO sales_contract
                      (order_line_id, contract_signed_date, sales_contract_no, contract_value,
                       performance_period, unsigned_contract_amount)
                    VALUES
                      (:order_line_id, :contract_signed_date, :sales_contract_no, :contract_value,
                       :performance_period, :unsigned_contract_amount)
                    """
                ),
                {
                    "order_line_id": order_line_id,
                    "contract_signed_date": _as_date(_row_value(row, 66)),
                    "sales_contract_no": _as_text(_row_value(row, 67)),
                    "contract_value": _as_decimal(_row_value(row, 68)),
                    "performance_period": _as_text(_row_value(row, 69)),
                    "unsigned_contract_amount": _as_decimal(_row_value(row, 70)),
                },
            )

            conn.execute(
                text(
                    """
                    INSERT INTO sales_invoice
                      (order_line_id, phase_no, invoice_doc_no, invoice_date, invoice_date_text,
                       invoice_no, invoice_amount, pending_invoice_amount, delivered_not_invoiced_amount)
                    VALUES
                      (:order_line_id, 1, :invoice_doc_no, :invoice_date, :invoice_date_text,
                       :invoice_no, :invoice_amount, :pending_invoice_amount, :delivered_not_invoiced_amount)
                    """
                ),
                {
                    "order_line_id": order_line_id,
                    "invoice_doc_no": _as_text(_row_value(row, 71)),
                    "invoice_date": _as_date(_row_value(row, 72)),
                    "invoice_date_text": _as_text(_row_value(row, 72)),
                    "invoice_no": _as_text(_row_value(row, 73)),
                    "invoice_amount": _as_decimal(_row_value(row, 74)),
                    "pending_invoice_amount": _as_decimal(_row_value(row, 75)),
                    "delivered_not_invoiced_amount": _as_decimal(_row_value(row, 76)),
                },
            )

            _insert_receipt(conn, order_line_id, 1, row, 77, 78, 79, 80)
            _insert_receipt(conn, order_line_id, 2, row, 81, 82, 83, 84)

            success_rows += 1
        except Exception as exc:  # noqa: BLE001 - record row-level import issue and continue
            failed_rows += 1
            conn.execute(
                text(
                    """
                    INSERT INTO ledger_raw_row
                      (import_batch_id, excel_row_no, row_hash, raw_json, parse_status, parse_message, project_code, order_no)
                    VALUES
                      (:batch_id, :excel_row_no, :row_hash, CAST(:raw_json AS JSON), 'failed', :message, :project_code, :order_no)
                    ON DUPLICATE KEY UPDATE parse_status = 'failed', parse_message = VALUES(parse_message)
                    """
                ),
                {
                    "batch_id": batch_id,
                    "excel_row_no": excel_row_no,
                    "row_hash": row_hash,
                    "raw_json": raw_json,
                    "message": str(exc)[:1000],
                    "project_code": project_code,
                    "order_no": order_no,
                },
            )

    conn.execute(
        text(
            """
            UPDATE import_batch
            SET total_rows = :total_rows,
                success_rows = :success_rows,
                failed_rows = :failed_rows,
                status = :status
            WHERE id = :batch_id
            """
        ),
        {
            "batch_id": batch_id,
            "total_rows": success_rows + failed_rows,
            "success_rows": success_rows,
            "failed_rows": failed_rows,
            "status": "completed" if failed_rows == 0 else "failed",
        },
    )

    conn.execute(
        text(
            """
            INSERT INTO operation_log (user_name, module_name, action_name, detail, status)
            VALUES ('system', '数据导入', 'import_excel', :detail, :status)
            """
        ),
        {
            "detail": f"导入 {workbook_path.name}: 成功 {success_rows} 行，失败 {failed_rows} 行",
            "status": "success" if failed_rows == 0 else "failed",
        },
    )

    conn.execute(
        text(
            """
            INSERT INTO backup_record (file_name, file_size_label, backup_type, status, storage_path)
            VALUES ('initial_import_snapshot.sql.gz', '0 MB', 'manual', 'success', 'local')
            """
        )
    )

    return {
        "batch_id": batch_id,
        "source_file": workbook_path.name,
        "success_rows": success_rows,
        "failed_rows": failed_rows,
    }


def _insert_phase(conn: Connection, table: str, order_line_id: int, phase_no: int, row: tuple[Any, ...], date_pos: int, code_pos: int, amount_pos: int) -> None:
    value = _as_decimal(_row_value(row, amount_pos))
    if value is None and _row_value(row, date_pos) is None and _row_value(row, code_pos) is None:
        return
    if table == "purchase_invoice":
        conn.execute(
            text(
                """
                INSERT INTO purchase_invoice
                  (order_line_id, phase_no, received_invoice_date, received_invoice_date_text, invoice_no, invoice_amount)
                VALUES
                  (:order_line_id, :phase_no, :date_value, :date_text, :code_value, :amount_value)
                """
            ),
            {
                "order_line_id": order_line_id,
                "phase_no": phase_no,
                "date_value": _as_date(_row_value(row, date_pos)),
                "date_text": _as_text(_row_value(row, date_pos)),
                "code_value": _as_text(_row_value(row, code_pos)),
                "amount_value": value,
            },
        )
    elif table == "warehouse_entry":
        conn.execute(
            text(
                """
                INSERT INTO warehouse_entry
                  (order_line_id, phase_no, warehouse_date, warehouse_date_text, voucher_no, warehouse_amount)
                VALUES
                  (:order_line_id, :phase_no, :date_value, :date_text, :code_value, :amount_value)
                """
            ),
            {
                "order_line_id": order_line_id,
                "phase_no": phase_no,
                "date_value": _as_date(_row_value(row, date_pos)),
                "date_text": _as_text(_row_value(row, date_pos)),
                "code_value": _as_text(_row_value(row, code_pos)),
                "amount_value": value,
            },
        )


def _insert_payment(conn: Connection, order_line_id: int, phase_no: int, due_date: date | None, payment_date: date | None, voucher_no: str | None, amount: float | None) -> None:
    if due_date is None and payment_date is None and voucher_no is None and amount is None:
        return
    conn.execute(
        text(
            """
            INSERT INTO purchase_payment
              (order_line_id, phase_no, due_payment_date, payment_date, payment_date_text, payment_voucher_no, payment_amount)
            VALUES
              (:order_line_id, :phase_no, :due_date, :payment_date, :payment_date_text, :voucher_no, :amount)
            """
        ),
        {
            "order_line_id": order_line_id,
            "phase_no": phase_no,
            "due_date": due_date,
            "payment_date": payment_date,
            "payment_date_text": payment_date.isoformat() if payment_date else None,
            "voucher_no": voucher_no,
            "amount": amount,
        },
    )


def _insert_receipt(conn: Connection, order_line_id: int, phase_no: int, row: tuple[Any, ...], date_pos: int, notice_pos: int, amount_pos: int, ratio_pos: int) -> None:
    amount = _as_decimal(_row_value(row, amount_pos))
    if amount is None and _row_value(row, date_pos) is None and _row_value(row, notice_pos) is None:
        return
    conn.execute(
        text(
            """
            INSERT INTO sales_receipt
              (order_line_id, phase_no, receipt_date, receipt_date_text, payment_notice_no, receipt_amount, receipt_ratio)
            VALUES
              (:order_line_id, :phase_no, :receipt_date, :receipt_date_text, :payment_notice_no, :receipt_amount, :receipt_ratio)
            """
        ),
        {
            "order_line_id": order_line_id,
            "phase_no": phase_no,
            "receipt_date": _as_date(_row_value(row, date_pos)),
            "receipt_date_text": _as_text(_row_value(row, date_pos)),
            "payment_notice_no": _as_text(_row_value(row, notice_pos)),
            "receipt_amount": amount,
            "receipt_ratio": _as_decimal(_row_value(row, ratio_pos)),
        },
    )
